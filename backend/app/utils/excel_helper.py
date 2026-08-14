"""
Excel 导入导出工具类（基于 openpyxl）
实现功能：
1. 批量用户/项目数据导入
2. 项目名单/经费台账/评审结果/操作日志导出
3. 导入模板下载
4. 导入错误校验（逐行校验 + 错误信息汇总）
"""
import os
import io
import re
from typing import List, Dict, Any, Tuple, Optional, Type, Callable
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore


from app.core.config import settings
from app.core.exceptions import ParamValidateException


# ====================================================================
# 列定义：用于生成模板/读取/校验
# ====================================================================
class ColumnDef:
    """单列表格定义"""
    def __init__(
        self, header: str, field: str, col_type: Type = str,
        required: bool = False, max_length: Optional[int] = None,
        choices: Optional[Dict[Any, Any]] = None,  # 显示值 -> 实际值 映射
        pattern: Optional[str] = None,             # 正则校验
        default: Any = None,
        desc: str = "",
    ):
        self.header = header
        self.field = field
        self.col_type = col_type
        self.required = required
        self.max_length = max_length
        self.choices = choices
        self.pattern = re.compile(pattern) if pattern else None
        self.default = default
        self.desc = desc

    def validate(self, raw_value: Any) -> Tuple[Any, Optional[str]]:
        """
        校验单元格原始值，返回 (转换后值, 错误信息或None)
        """
        # 1. 必填校验
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            if self.required:
                return None, f"必填项为空"
            return self.default, None

        # 字符串预处理
        value = raw_value
        if isinstance(value, str):
            value = value.strip()

        # 2. 枚举选择
        if self.choices:
            if value not in self.choices:
                return None, f"值 '{value}' 不在允许选项中: {list(self.choices.keys())}"
            value = self.choices[value]

        # 3. 类型转换
        try:
            if self.col_type == bool:
                if isinstance(value, bool):
                    pass
                elif isinstance(value, (int, float)):
                    value = bool(value)
                else:
                    value = str(value).lower() in ("true", "1", "是", "y", "yes")
            elif self.col_type == int:
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                else:
                    value = int(value)
            elif self.col_type == float:
                value = float(value)
            elif self.col_type == Decimal:
                value = Decimal(str(value))
            elif self.col_type == date:
                if isinstance(value, datetime):
                    value = value.date()
                elif isinstance(value, date):
                    pass
                else:
                    # 常见格式尝试
                    text = str(value)
                    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
                        try:
                            value = datetime.strptime(text, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        return None, f"日期格式无效，支持 YYYY-MM-DD"
            elif self.col_type == datetime:
                if isinstance(value, datetime):
                    pass
                else:
                    text = str(value)
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                        try:
                            value = datetime.strptime(text, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        return None, f"时间格式无效"
            elif self.col_type == str:
                value = str(value)
        except (ValueError, InvalidOperation):
            return None, f"类型转换失败，期望 {self.col_type.__name__}"

        # 4. 长度校验
        if self.max_length and isinstance(value, str) and len(value) > self.max_length:
            return None, f"长度超过最大限制 {self.max_length}"

        # 5. 正则校验
        if self.pattern and isinstance(value, str) and not self.pattern.match(value):
            return None, f"格式不匹配规则"

        return value, None


# ====================================================================
# 模板定义：用户导入 / 项目导入
# ====================================================================
USER_IMPORT_COLUMNS: List[ColumnDef] = [
    ColumnDef("登录账号*", "username", str, required=True, max_length=64,
              pattern=r"^[A-Za-z0-9_]{2,64}$", desc="字母数字下划线"),
    ColumnDef("姓名*", "real_name", str, required=True, max_length=64),
    ColumnDef("密码*", "password", str, required=True, max_length=64),
    ColumnDef("邮箱", "email", str, max_length=128),
    ColumnDef("手机号", "phone", str, max_length=20, pattern=r"^1[3-9]\d{9}$"),
    ColumnDef("角色*", "role", int, required=True,
              choices={"学生": 1, "指导教师": 2, "评审专家": 3, "系统管理员": 4}),
    ColumnDef("学院编码*", "college_id", int, required=True,
              desc="请使用学院ID；可在学院列表中查询"),
    ColumnDef("状态", "status", int, default=1, choices={"启用": 1, "禁用": 0}),
]

PROJECT_IMPORT_COLUMNS: List[ColumnDef] = [
    ColumnDef("项目名称*", "project_name", str, required=True, max_length=255),
    ColumnDef("项目类型*", "project_type", int, required=True,
              choices={"创新训练": 1, "创业训练": 2, "创业实践": 3}),
    ColumnDef("项目级别*", "project_level", int, required=True,
              choices={"校级": 1, "省级": 2, "国家级": 3}),
    ColumnDef("学院编码*", "college_id", int, required=True),
    ColumnDef("负责人学号*", "leader_username", str, required=True,
              desc="系统中必须存在的学生账号"),
    ColumnDef("指导教师工号", "teacher_username", str, max_length=64),
    ColumnDef("开始日期", "start_date", date),
    ColumnDef("结束日期", "end_date", date),
    ColumnDef("预算总额", "budget_amount", Decimal, default=Decimal("0")),
    ColumnDef("项目简介", "project_summary", str, max_length=5000),
]


# ====================================================================
# Excel 工具类
# ====================================================================
class ExcelHelper:
    """Excel 导入导出工具类"""

    @staticmethod
    def _ensure_openpyxl():
        if Workbook is None:
            raise RuntimeError("请先安装 openpyxl: pip install openpyxl")

    # ---------- 生成模板 ----------
    @staticmethod
    def generate_template(columns: List[ColumnDef], sheet_name: str = "Sheet1",
                          with_comments: bool = True) -> bytes:
        """
        根据列定义生成空白导入模板（含表头、下拉注释、示例行）
        :returns: xlsx 文件字节流
        """
        ExcelHelper._ensure_openpyxl()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 1. 表头
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # 2. 注释行（第2行，说明）
        if with_comments:
            for col_idx, col in enumerate(columns, start=1):
                parts = []
                if col.desc:
                    parts.append(col.desc)
                if col.choices:
                    parts.append("可选: " + "/".join(str(k) for k in col.choices.keys()))
                parts.append("必填" if col.required else "选填")
                cell = ws.cell(row=2, column=col_idx, value=" | ".join(parts))
                cell.font = Font(italic=True, color="666666", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = thin_border

        # 3. 列宽自适应
        for col_idx, col in enumerate(columns, start=1):
            width = max(12, min(50, len(col.header) * 2 + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # 冻结表头
        ws.freeze_panes = "A3" if with_comments else "A2"

        # 写内存返回
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ---------- 数据解析 ----------
    @staticmethod
    def parse_workbook(file_bytes: bytes, columns: List[ColumnDef],
                       header_row: int = 1, data_start_row: int = 3) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        解析上传的Excel文件为数据字典列表 + 错误信息
        :param file_bytes: 上传文件的bytes
        :param columns: 列定义（顺序必须与Excel列顺序一致）
        :param header_row: 表头所在行号
        :param data_start_row: 数据起始行（含），默认跳过注释行
        :returns: (成功解析的数据列表, 错误信息列表)
        """
        ExcelHelper._ensure_openpyxl()
        errors: List[str] = []
        results: List[Dict[str, Any]] = []
        try:
            buffer = io.BytesIO(file_bytes)
            wb = load_workbook(buffer, data_only=True, read_only=False)
            ws = wb.active
        except Exception as e:
            return [], [f"Excel文件读取失败: {e}"]

        # 表头校验
        headers = [ws.cell(row=header_row, column=i).value for i in range(1, len(columns) + 1)]
        expected_headers = [c.header for c in columns]
        for idx, (h, exp) in enumerate(zip(headers, expected_headers)):
            # 去掉必填标记*对比
            norm_h = (str(h) if h else "").replace("*", "").strip()
            norm_exp = exp.replace("*", "").strip()
            if norm_h != norm_exp:
                errors.append(f"第{header_row}行第{idx + 1}列表头不匹配：期望'{exp}'，实际'{h}'")
        if errors:
            return [], errors

        # 逐行解析
        max_row = ws.max_row or 0
        if max_row < data_start_row:
            return [], ["未检测到有效数据行"]

        for row_idx in range(data_start_row, max_row + 1):
            row_data: Dict[str, Any] = {}
            row_errors: List[str] = []
            all_empty = True
            for col_idx, col_def in enumerate(columns, start=1):
                raw = ws.cell(row=row_idx, column=col_idx).value
                value, err = col_def.validate(raw)
                if raw is not None and not (isinstance(raw, str) and raw.strip() == ""):
                    all_empty = False
                if err:
                    row_errors.append(f"列[{col_def.header}]: {err}")
                else:
                    row_data[col_def.field] = value
            if all_empty:
                continue  # 空行跳过
            if row_errors:
                errors.append(f"第{row_idx}行: " + "；".join(row_errors))
            else:
                results.append(row_data)
        return results, errors

    # ---------- 导出列表 ----------
    @staticmethod
    def export_list(
        columns: List[Tuple[str, str, Optional[Callable]]],  # [(表头, 字段或key, 格式化函数可选)]
        data_rows: List[Dict[str, Any]],
        sheet_name: str = "Sheet1",
    ) -> bytes:
        """
        将数据列表导出为Excel二进制
        columns 示例: [("项目名称", "project_name"), ("预算", "budget_amount", lambda v: f"¥{v:.2f}")]
        """
        ExcelHelper._ensure_openpyxl()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 写表头
        for col_idx, (header, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) * 2 + 2)

        # 写数据
        for r_idx, row in enumerate(data_rows, start=2):
            for c_idx, (_, field, fmt) in enumerate(columns, start=1):
                val = row.get(field) if isinstance(row, dict) else getattr(row, field, None)
                if fmt and val is not None:
                    try:
                        val = fmt(val)
                    except Exception:
                        pass
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
        ws.freeze_panes = "A2"
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


# ====================================================================
# 常用导出：项目名单 / 经费台账 / 评审结果 / 操作日志 / 用户列表
# ====================================================================
class ExportTemplates:
    """预定义常用导出列，配合 export_list 使用"""

    @staticmethod
    def project_list_columns():
        return [
            ("项目编号", "project_no", None),
            ("项目名称", "project_name", None),
            ("项目类型", "project_type_name", None),
            ("项目级别", "project_level_name", None),
            ("所属学院", "college_name", None),
            ("负责人", "leader_name", None),
            ("指导教师", "teacher_name", None),
            ("开始日期", "start_date", lambda v: str(v) if v else ""),
            ("结束日期", "end_date", lambda v: str(v) if v else ""),
            ("预算(元)", "budget_amount", lambda v: f"{Decimal(str(v or 0)):.2f}"),
            ("已用(元)", "used_amount", lambda v: f"{Decimal(str(v or 0)):.2f}"),
            ("项目状态", "status_name", None),
            ("提交时间", "submit_time", lambda v: v.strftime("%Y-%m-%d %H:%M") if v else ""),
        ]

    @staticmethod
    def user_list_columns():
        return [
            ("账号", "username", None),
            ("姓名", "real_name", None),
            ("角色", "role_name", None),
            ("学院", "college_name", None),
            ("邮箱", "email", None),
            ("手机号", "phone", None),
            ("状态", "status", lambda v: "启用" if v == 1 else "禁用"),
            ("创建时间", "created_at", lambda v: v.strftime("%Y-%m-%d %H:%M") if v else ""),
        ]

    @staticmethod
    def operation_log_columns():
        return [
            ("操作时间", "operation_time", lambda v: v.strftime("%Y-%m-%d %H:%M:%S") if v else ""),
            ("操作人", "real_name", None),
            ("账号", "username", None),
            ("模块", "module_name", None),
            ("操作类型", "operation_type", None),
            ("操作描述", "operation_desc", None),
            ("请求方式", "request_method", None),
            ("请求URL", "request_url", None),
            ("IP地址", "ip_address", None),
            ("耗时(ms)", "cost_time", None),
            ("响应码", "response_code", None),
        ]

    @staticmethod
    def review_result_columns():
        return [
            ("项目编号", "project_no", None),
            ("项目名称", "project_name", None),
            ("学院", "college_name", None),
            ("评审阶段", "review_stage_name", None),
            ("评审专家", "reviewer_name", None),
            ("评分", "score", lambda v: f"{v}" if v is not None else ""),
            ("评审结果", "review_result_name", None),
            ("评审意见", "review_comment", None),
            ("评审时间", "review_time", lambda v: v.strftime("%Y-%m-%d %H:%M") if v else ""),
        ]
